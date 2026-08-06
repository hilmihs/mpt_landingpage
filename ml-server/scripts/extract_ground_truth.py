"""
Ambil ground truth penilaian Ustadzah dari spreadsheet "Penilaian Hasil Rekaman
Al Faatihah Mustahik" → JSON siap pakai untuk eval_ground_truth.py.

KENAPA ADA
Sampai sekarang tidak ada satu pun angka yang bisa dipakai menilai akurasi
mesin: pengajar baru mulai memakai instrumen per-segmen, dan pasangannya masih
nol. Spreadsheet ini memuat 789 rekaman yang sudah dihitung jumlah lahn jaliy
dan khafiy-nya oleh Ustadzah sejak Oktober 2025 — ground truth yang jauh lebih
besar dan lebih tua daripada apa pun yang bisa dikumpulkan dari aplikasi.

Angkanya TOTAL per rekaman, bukan per segmen. Jadi ini tidak bisa dipakai
menghitung skor 1-10, dan memang bukan itu tujuannya. Yang diukur adalah hal
yang lebih mendasar: apakah mesin melihat jumlah kesalahan yang kira-kira sama
dengan yang dilihat manusia.

DATA PRIBADI
Spreadsheet memuat nama lengkap dan nomor WhatsApp. Keduanya TIDAK ikut keluar
dari sini. Yang disimpan hanya id berkas Drive, dua hitungan, dan jenis kelamin
— cukup untuk mengukur akurasi, tidak cukup untuk mengenali orang.

CARA PAKAI
    # unduh spreadsheet sebagai .xlsx dari Google Sheets, lalu:
    python scripts/extract_ground_truth.py rekaman.xlsx -o ground_truth.json
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import warnings

SHEET_NAME = "Rekaman Mustahik"

# Kolom pada sheet (1-based). Header ada di baris 1, data mulai baris 2.
COL_GENDER = 8   # H
COL_BERKAS = 9   # I — teksnya nama berkas, hyperlink-nya menuju Drive
COL_JALIY = 10   # J
COL_KHAFIY = 11  # K

_DRIVE_ID = re.compile(r"/d/([A-Za-z0-9_-]{20,})")


def extract(path: str) -> list[dict]:
    warnings.filterwarnings("ignore")  # openpyxl mengeluh soal pivot cache
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("openpyxl belum terpasang: pip install openpyxl")

    wb = load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"tab {SHEET_NAME!r} tidak ada. Yang tersedia: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        jaliy = ws.cell(r, COL_JALIY).value
        khafiy = ws.cell(r, COL_KHAFIY).value
        # Baris tanpa kedua angka belum dinilai — bukan nol, tapi belum ada.
        if not isinstance(jaliy, (int, float)) or not isinstance(khafiy, (int, float)):
            continue

        sel = ws.cell(r, COL_BERKAS)
        link = sel.hyperlink.target if getattr(sel, "hyperlink", None) else None
        m = _DRIVE_ID.search(link or "")
        if not m:
            continue

        rows.append(
            {
                "drive_id": m.group(1),
                "jaliy": int(jaliy),
                "khafiy": int(khafiy),
                "gender": str(ws.cell(r, COL_GENDER).value or "").strip().lower(),
                # Nomor baris saja, supaya temuan aneh bisa ditelusuri balik ke
                # spreadsheet tanpa menyimpan identitas siapa pun.
                "baris": r,
            }
        )
    return rows


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", help="spreadsheet yang diunduh sebagai .xlsx")
    ap.add_argument("-o", "--out", default="ground_truth.json")
    args = ap.parse_args()

    rows = extract(args.xlsx)
    if not rows:
        sys.exit("tidak ada baris yang punya jaliy, khafiy, DAN tautan Drive sekaligus")

    with open(args.out, "w", encoding="utf8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=1)

    unik = len({r["drive_id"] for r in rows})
    print(f"{len(rows)} baris ditulis ke {args.out} ({unik} berkas unik)")
    if unik != len(rows):
        print(f"  catatan: {len(rows) - unik} baris menunjuk berkas yang sama")


if __name__ == "__main__":
    main()
